# Установлены: sqlalchemy, openpyxl (работа с excel)

from sqlite3 import connect
from xml.sax.handler import DTDHandler
from openpyxl import Workbook, utils, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import Border, Side

from glob import glob
# Нужно, чтобы удалять файлы
import os

from config_main import QUERY_OUTPUT_FILES_PATH, EXTENDED_RESULT, OUTPUT_FILES_PATH, RESULT_FILE
from config_sections import REPORT_SECTIONS
import functions
from functions import session
from model import InitialData

# Функция сохраняет заданный запрос в файл
# Получает:
# - query - сам запрос
# - file_path - путь к файлу
# - file_name - имя файла
# - list_title - заголовок листа в xlsx файле
# (это будет использовано первой строчкой листа,
# также начало заголовка до первого пробела будет
# использовано как название листа)
def query_to_xlsx(query, file_path, file_name, list_title):
    result_file_query = f"{file_path}{file_name}.xlsx"
    list_name = list_title[:list_title.find(" ")]
    # Если файл есть, используем его, если нет,
    # то создаем новый
    if os.path.isfile(result_file_query):
        file_exist = True
    else:
        file_exist = False
    if file_exist:
        wbq = load_workbook(result_file_query)
        wsq = wbq.create_sheet(list_name)
    else:
        wbq = Workbook()
        wsq = wbq.active
        wsq.title = list_name
    _ = wsq.cell(column=1, 
            row=1, 
            value=list_title)
    for row, query_string in enumerate(query):
        for col, query_string_value in enumerate(query_string):
            _ = wsq.cell(column=col+1, 
            row=row+2, 
            value=query_string_value)
    # TODO - одинаковые операторы в условии - похоже на ошибку
    if file_exist:
        wbq.save(filename = result_file_query)
    else:
        wbq.save(filename = result_file_query)


def fill_cell(ws, section, worksheet_need, return_name):
    '''
    Занести значение в ячейку
    '''
    if return_name == None and worksheet_need:
        _ = ws.cell(
            column=col + 2,
            row=section["row_for_excel"],
            value=section["function_name"](col, ws, REPORT_SECTIONS)
        )
        return
    if return_name == None:
        _ = ws.cell(
            column=col + 2,
            row=section["row_for_excel"],
            value=section["function_name"](period_names[col].period_name)
        )
        return
    if worksheet_need:
        _ = ws.cell(
            column=col + 2,
            row=section["row_for_excel"],
            value=section["function_name"](col, ws, REPORT_SECTIONS)[return_name]
        )
        return
    function_return = section["function_name"](period_names[col].period_name)
    _ = ws.cell(
        column=col + 2,
        row=section["row_for_excel"],
        value=function_return[return_name]
    )
    if EXTENDED_RESULT and section["query_returns_name"] != None:
        _ = query_to_xlsx(
            function_return[section["query_returns_name"]],
            QUERY_OUTPUT_FILES_PATH,
            str(period_names[col].period_name)[:7],
            section["section_id"]
        )


def add_thin_borders(ws, len_period_names):
    '''
    Добавить тонкие рамки в таблицу
    '''
    # Ищем максимальный номер выводимой строки
    max_row = max(map(
        lambda section: section["row_for_excel"],
        filter(lambda section: section["output_to_report"], REPORT_SECTIONS)
    ))
    # Настраиваем тип рамки
    thins = Side(border_style="thin", color="000000")
    # "Рисуем" рамки
    for col in range(len_period_names + 1):
        for row in range(max_row):
            ws[
                f"{utils.cell.get_column_letter(col + 1)}{row+1}"
            ].border = Border(
                top=thins,
                bottom=thins,
                left=thins,
                right=thins
            )


def setting_width_of_columns(ws, len_period_names):
    '''
    Зададим ширины столбцов
    '''
    # для первого столбца увеличим ширину до 88
    ws.column_dimensions[utils.cell.get_column_letter(1)].width = 88
    # для остальных столбцов зададим ширину 14
    for col in range(len_period_names):
        ws.column_dimensions[utils.cell.get_column_letter(col+2)].width = 14


def format_period_names(ws, len_period_names):
    '''
    Установить правильное форматирование ячеек с названиями периодов
    '''
    for col in range(len_period_names):
        # форматирование строки периодов
        ws[
            f"{utils.cell.get_column_letter(col+2)}{1}"
        ].number_format = "mmmm yyyy;@"



def main():
    result_file = f"{OUTPUT_FILES_PATH}{RESULT_FILE}"
    print(f"Файл: {result_file}")

    # Если предполагается расширенный вывод, то очистим папку
    # для дополнительных файлов
    # (стираем все xlsx файлы в ней)
    if EXTENDED_RESULT:
        list_of_xlsx = glob(f"{QUERY_OUTPUT_FILES_PATH}*.xlsx")
        for file_for_delete in list_of_xlsx:
            os.remove(file_for_delete)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный отчёт"

    # Выведем в excel названия периодов
    period_names = session.query(InitialData).all()
    # Сделаю отдельную переменную, чтобы каждый раз не считать
    len_period_names = len(period_names)
    print(f"Сколько периодов: {len_period_names}")

    # выполним из REPORT_SECTIONS функции, где "output_to_report": True
    for section in filter(lambda section: section["output_to_report"], REPORT_SECTIONS):
        print(f"-> ВЫВОДИМ СЕКЦИЮ: {section['section_id']}")
        # !!! - для отладки
        # wb.save(filename = result_file)
        print("!!! - отрабатываем секцию - !!!")
        print(section)
        _ = ws.cell(
            column=1,
            row=section["row_for_excel"],
            value=section["section_name_for_report"]
        )
        for col in range(
            len_period_names -
            (lambda last_period: 0 if last_period else 1)(
            section["output_the_last_period"])
        ):
            fill_cell(
                ws,
                section,
                worksheet_need=section["worksheet_need"],
                return_name=section["functon_returns_name"]
            )
            # Установим правильное форматирование ячеек
            # для строк c "money_format" == True
            # с 2 знаками после запятой и разделение троек разрядов пробелами
            # Нужный формат BUILTIN_FORMATS[4] (выглядит как '#,##0.00'))
            # В русском экселе будет виден как '# ##0,00'
            if section["money_format"]:
                ws[
                    f"{utils.cell.get_column_letter(col+2)}{section['row_for_excel']}"
                ].number_format = BUILTIN_FORMATS[4]

    # Зададим ширины столбцов
    setting_width_of_columns(ws, len_period_names)

    # Установим правильное форматирование ячеек строки названий периодов
    format_period_names(ws, len_period_names)

    # Добавить рамочки
    add_thin_borders(ws, len_period_names)

    # Зафиксируем строку A и столбец 1 (заголовки)
    ws.freeze_panes = "B2"

    wb.save(filename=result_file)


if __name__ == "__main__":
    main()
